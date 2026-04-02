"""
generate_op.py — Streamlit app
Tab 1: Upload a ZIP of resumes → match against job-email vectors → download Excel.
Tab 2: Paste a job requirement → find matching LinkedIn profiles via Google.
"""

import json
import logging
import re
import tempfile
import time
import zipfile
from io import BytesIO
from pathlib import Path
from urllib.parse import quote_plus

import numpy as np
import streamlit as st
from googlesearch import search as google_search
from sentence_transformers import SentenceTransformer

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

VECTOR_DIR = Path("vectors")
EMBED_MODEL_NAME = "all-MiniLM-L6-v2"

TECH_SKILLS = {
    "python", "java", "javascript", "typescript", "c#", "c++", "go", "golang",
    "ruby", "rust", "scala", "kotlin", "swift", "php", "perl", "r",
    "sql", "nosql", "mysql", "postgresql", "postgres", "oracle", "mongodb",
    "cassandra", "redis", "elasticsearch", "dynamodb", "sqlite", "snowflake",
    "aws", "azure", "gcp", "google cloud", "ec2", "s3", "lambda", "ecs",
    "eks", "fargate", "cloudformation", "terraform", "ansible", "pulumi",
    "docker", "kubernetes", "k8s", "openshift", "helm", "istio",
    "jenkins", "gitlab", "github actions", "ci/cd", "circleci", "bamboo",
    "react", "angular", "vue", "nextjs", "next.js", "nuxt", "svelte",
    "node", "nodejs", "node.js", "express", "fastapi", "flask", "django",
    "spring", "springboot", "spring boot", ".net", "dotnet", "asp.net",
    "html", "css", "sass", "tailwind", "bootstrap",
    "rest", "restful", "graphql", "grpc", "soap", "api",
    "kafka", "rabbitmq", "sqs", "sns", "activemq", "celery",
    "spark", "pyspark", "hadoop", "airflow", "databricks", "delta lake",
    "dbt", "etl", "elt", "data pipeline", "data engineering",
    "pandas", "numpy", "scikit-learn", "tensorflow", "pytorch", "keras",
    "machine learning", "deep learning", "nlp", "computer vision", "llm",
    "tableau", "power bi", "looker", "grafana",
    "git", "jira", "confluence", "agile", "scrum",
    "linux", "unix", "bash", "powershell", "shell scripting",
    "microservices", "serverless", "event-driven",
    "oauth", "jwt", "saml", "sso", "ldap",
    "selenium", "cypress", "playwright", "junit", "pytest", "jest",
    "hive", "pig", "presto", "athena", "redshift", "bigquery",
    "salesforce", "sap", "servicenow", "pega", "appian",
    "splunk", "datadog", "new relic", "prometheus", "elk",
    "figma", "sketch",
    "azure devops", "azure data factory", "adls", "glue", "step functions",
    "unity catalog", "medallion", "star schema", "snowflake schema",
    "servicenow", "informatica", "talend", "nifi", "fivetran",
}


def _extract_email(sender: str) -> str:
    match = re.search(r"<([^>]+)>", sender)
    if match:
        return match.group(1)
    if "@" in sender:
        return sender.strip()
    return sender


def extract_skills(text: str) -> set[str]:
    text_lower = text.lower()
    found = set()
    for skill in TECH_SKILLS:
        if re.search(rf"\b{re.escape(skill)}\b", text_lower):
            found.add(skill)
    return found


# ---------------------------------------------------------------------------
# Resume parsing
# ---------------------------------------------------------------------------

def _text_from_pdf(file_bytes: bytes) -> str:
    import fitz
    parts = []
    with fitz.open(stream=file_bytes, filetype="pdf") as doc:
        for page in doc:
            parts.append(page.get_text())
    return "\n".join(parts)


def _text_from_docx(file_bytes: bytes) -> str:
    import docx
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    doc = docx.Document(tmp_path)
    return "\n".join(p.text for p in doc.paragraphs)


def parse_resume_bytes(file_bytes: bytes, filename: str) -> str:
    name = filename.lower()
    if name.endswith(".pdf"):
        return _text_from_pdf(file_bytes)
    elif name.endswith(".docx"):
        return _text_from_docx(file_bytes)
    return file_bytes.decode("utf-8", errors="ignore")


# ---------------------------------------------------------------------------
# ZIP extraction — returns {folder_name: [(resume_name, text), ...]}
# ---------------------------------------------------------------------------

def extract_zip(uploaded_zip) -> dict[str, list[tuple[str, str]]]:
    """Parse a ZIP file into {tech_stack_folder: [(filename, resume_text), ...]}."""
    raw = uploaded_zip.read()
    result: dict[str, list[tuple[str, str]]] = {}

    with zipfile.ZipFile(BytesIO(raw)) as zf:
        for entry in zf.namelist():
            if entry.endswith("/"):
                continue
            parts = Path(entry).parts
            if len(parts) < 2:
                folder = "General"
                fname = parts[0]
            else:
                folder = parts[-2]
                fname = parts[-1]

            if not (fname.lower().endswith(".pdf") or fname.lower().endswith(".docx")):
                log.info("Skipping non-resume file: %s", entry)
                continue

            file_bytes = zf.read(entry)
            try:
                text = parse_resume_bytes(file_bytes, fname)
            except Exception as e:
                log.warning("Could not parse %s: %s", entry, e)
                continue

            result.setdefault(folder, []).append((fname, text))

    log.info("ZIP extracted — %d folders, %d total resumes",
             len(result), sum(len(v) for v in result.values()))
    return result


# ---------------------------------------------------------------------------
# Vector helpers
# ---------------------------------------------------------------------------

def load_all_vectors() -> list[dict]:
    records = []
    if not VECTOR_DIR.exists():
        return records
    for jf in sorted(VECTOR_DIR.glob("*.jsonl")):
        with jf.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    records.append(json.loads(line))
    log.info("Loaded %d total vectors from %s", len(records), VECTOR_DIR)
    return records


def cosine_similarity(a: np.ndarray, b: np.ndarray) -> float:
    dot = np.dot(a, b)
    norm = np.linalg.norm(a) * np.linalg.norm(b)
    if norm == 0:
        return 0.0
    return float(dot / norm)


def match_resumes_to_jobs(
    resume_data: list[tuple[str, np.ndarray, set[str]]],
    records: list[dict],
) -> list[dict]:
    """
    For each job email, compute similarity with every resume.
    Return rows sorted by best similarity descending.
    Each row includes which resume(s) matched and their scores.
    """
    rows = []
    for rec in records:
        job_vec = np.array(rec["vector"], dtype=np.float32)
        job_text = f"{rec.get('subject', '')} {rec.get('body', '')}"
        job_skills = extract_skills(job_text)

        per_resume = []
        best_score = 0.0
        for resume_name, resume_vec, resume_skills in resume_data:
            score = cosine_similarity(resume_vec, job_vec)
            missing = sorted(job_skills - resume_skills)
            per_resume.append((resume_name, round(score * 100, 2), missing))
            best_score = max(best_score, score)

        per_resume.sort(key=lambda x: x[1], reverse=True)
        matching_names = ", ".join(f"{name} ({sc}%)" for name, sc, _ in per_resume)

        all_missing: dict[str, list[str]] = {}
        for name, _sc, miss in per_resume:
            all_missing[name] = miss

        missing_summary_parts = []
        for name, miss in all_missing.items():
            if miss:
                missing_summary_parts.append(f"{name}: {', '.join(miss)}")
        missing_summary = " | ".join(missing_summary_parts) if missing_summary_parts else "None"

        total = len(per_resume)
        buckets = {"90-100%": 0, "80-90%": 0, "70-80%": 0, "60-70%": 0, "Below 60%": 0}
        for _, sc, _ in per_resume:
            if sc >= 90:
                buckets["90-100%"] += 1
            elif sc >= 80:
                buckets["80-90%"] += 1
            elif sc >= 70:
                buckets["70-80%"] += 1
            elif sc >= 60:
                buckets["60-70%"] += 1
            else:
                buckets["Below 60%"] += 1
        bucket_summary = ", ".join(
            f"{count}/{total} ({label})" for label, count in buckets.items() if count > 0
        )

        top_resume_names = [name for name, sc, _ in per_resume if sc >= 60]
        if not top_resume_names:
            top_resume_names = [per_resume[0][0]] if per_resume else []

        sender_first = rec.get("sender_name", "").split()[0] if rec.get("sender_name", "").strip() else "Hi"
        subject_line = rec.get("subject", "the open position")

        if len(top_resume_names) == 1:
            candidate_text = f"a strong candidate ({top_resume_names[0]})"
        elif len(top_resume_names) > 1:
            candidate_text = f"{len(top_resume_names)} strong candidates ({', '.join(top_resume_names)})"
        else:
            candidate_text = "a candidate"

        draft_reply = (
            f"Hi {sender_first},\n\n"
            f"Thank you for sharing the requirement for \"{subject_line}\".\n\n"
            f"We have {candidate_text} who closely match your requirements. "
            f"Please find the attached resume(s) for your review.\n\n"
            f"Let us know a good time to discuss availability and next steps.\n\n"
            f"Best regards"
        )

        rows.append({
            "best_similarity": round(best_score * 100, 2),
            "subject": rec.get("subject", ""),
            "sender_name": rec.get("sender_name", ""),
            "sender_email": rec.get("sender_email", _extract_email(rec.get("sender", ""))),
            "status": rec.get("status", ""),
            "job_skills": ", ".join(sorted(job_skills)),
            "matching_resumes": matching_names,
            "resume_match_summary": bucket_summary,
            "missing_skills": missing_summary,
            "draft_reply": draft_reply,
            "body": rec.get("body", ""),
        })

    rows.sort(key=lambda x: x["best_similarity"], reverse=True)
    return rows


# ---------------------------------------------------------------------------
# Excel export — one sheet per tech-stack folder
# ---------------------------------------------------------------------------

def to_excel_bytes(all_results: dict[str, list[dict]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "Rank", "Best Similarity %", "Subject", "Sender Name", "Sender Email",
        "Status", "Job Skills", "Matching Resumes (score)", "Resume Match Summary",
        "Missing Skills per Resume", "Draft Reply", "Body",
    ]

    for folder_name, rows in all_results.items():
        sheet_name = folder_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

        for idx, row in enumerate(rows, start=1):
            ws.append([
                idx,
                row["best_similarity"],
                row["subject"],
                row["sender_name"],
                row["sender_email"],
                row["status"],
                row["job_skills"],
                row["matching_resumes"],
                row["resume_match_summary"],
                row["missing_skills"],
                row["draft_reply"],
                row["body"][:500],
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

@st.cache_resource
def load_model():
    return SentenceTransformer(EMBED_MODEL_NAME)


# ---------------------------------------------------------------------------
# LinkedIn profile search via Google
# ---------------------------------------------------------------------------

def build_linkedin_queries(role: str, skills: list[str], location: str) -> list[str]:
    """Build multiple Google queries from broad to narrow for better results."""
    queries = []

    # Query 1: role + top 3 skills
    q1_parts = ['site:linkedin.com/in']
    if role:
        q1_parts.append(f'"{role}"')
    for skill in skills[:3]:
        q1_parts.append(f'"{skill}"')
    if location and location.lower() != "remote":
        q1_parts.append(f'"{location}"')
    queries.append(" ".join(q1_parts))

    # Query 2: role + top 2 skills (broader)
    q2_parts = ['site:linkedin.com/in']
    if role:
        q2_parts.append(f'"{role}"')
    for skill in skills[:2]:
        q2_parts.append(skill)
    queries.append(" ".join(q2_parts))

    # Query 3: just role + unquoted skills (broadest)
    q3_parts = ['site:linkedin.com/in', role] + skills[:4]
    queries.append(" ".join(q3_parts))

    return queries


def parse_requirement(text: str) -> dict:
    """Extract role, skills, and location from pasted requirement text."""
    text_lower = text.lower()
    found_skills = []
    for skill in TECH_SKILLS:
        if re.search(rf"\b{re.escape(skill)}\b", text_lower):
            found_skills.append(skill)

    role = ""
    role_patterns = [
        r"(?:job\s*title|role|title|position)\s*[:;-]\s*(.+?)(?:\n|$)",
        r"(?:looking for|hiring|need|seeking)\s+(?:a|an)?\s*(.+?)(?:\.|,|\n|$)",
    ]
    for pat in role_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            role = m.group(1).strip()[:80]
            break

    location = ""
    loc_match = re.search(
        r"(?:location|loc|city)\s*[:;-]\s*(.+?)(?:\n|$)", text, re.IGNORECASE
    )
    if loc_match:
        location = loc_match.group(1).strip()

    return {"role": role, "skills": found_skills, "location": location}


def search_linkedin_profiles(queries: list[str], num_results: int = 15) -> tuple[list[dict], str]:
    """Try multiple queries from narrow to broad until we get results."""
    seen_urls = set()
    results = []
    used_query = ""

    for query in queries:
        log.info("Trying Google query: %s", query)
        used_query = query
        try:
            for url in google_search(query, num_results=num_results, lang="en"):
                if "linkedin.com/in/" not in url:
                    continue
                if url in seen_urls:
                    continue
                seen_urls.add(url)
                name_part = url.rstrip("/").split("/in/")[-1].split("?")[0]
                display_name = name_part.replace("-", " ").title()
                results.append({"name": display_name, "url": url})
                time.sleep(0.3)
        except Exception as e:
            log.warning("Google search error for query '%s': %s", query, e)

        if results:
            break
        log.info("No results, trying broader query...")
        time.sleep(1)

    return results, used_query


def profiles_to_excel(profiles: list[dict], query: str) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Profiles"
    headers = ["#", "Name", "LinkedIn URL", "Search Query"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    for idx, p in enumerate(profiles, 1):
        ws.append([idx, p["name"], p["url"], query if idx == 1 else ""])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

def tab_resume_matcher():
    st.header("Resume ↔ Job Email Matcher")
    st.write(
        "Upload a **ZIP file** with folders named by tech stack "
        "(e.g. `Java/`, `Python/`, `.NET/`), each containing resumes (PDF/DOCX). "
        "You'll get one Excel sheet per folder with ranked job matches + missing skills."
    )

    uploaded = st.file_uploader("Upload Resume ZIP", type=["zip"])

    if uploaded is not None:
        with st.spinner("Extracting ZIP..."):
            folder_resumes = extract_zip(uploaded)

        if not folder_resumes:
            st.error("No valid resumes (PDF/DOCX) found inside the ZIP.")
            return

        for folder, resumes in folder_resumes.items():
            st.success(f"**{folder}** — {len(resumes)} resume(s): {', '.join(r[0] for r in resumes)}")

        records = load_all_vectors()
        if not records:
            st.error("No job vectors found. Run `daily_vectorize.py` first.")
            return

        st.info(f"Loaded {len(records)} job email vectors from `{VECTOR_DIR}/`")

        model = load_model()
        all_results: dict[str, list[dict]] = {}

        progress = st.progress(0, text="Processing folders...")
        folder_names = list(folder_resumes.keys())

        for i, folder in enumerate(folder_names):
            resumes = folder_resumes[folder]
            log.info("Processing folder '%s' with %d resumes", folder, len(resumes))
            progress.progress((i) / len(folder_names), text=f"Processing {folder}...")

            resume_data = []
            for fname, text in resumes:
                vec = model.encode(text[:8000])
                skills = extract_skills(text)
                resume_data.append((fname, np.array(vec, dtype=np.float32), skills))

            matched = match_resumes_to_jobs(resume_data, records)
            all_results[folder] = matched

            st.subheader(f"{folder} — Top 10 matches")
            preview = []
            for r in matched[:10]:
                preview.append({
                    "Similarity %": r["best_similarity"],
                    "Subject": r["subject"],
                    "Sender Email": r["sender_email"],
                    "Resume Match Summary": r["resume_match_summary"],
                    "Matching Resumes": r["matching_resumes"],
                    "Missing Skills": r["missing_skills"][:100],
                })
            st.table(preview)

        progress.progress(1.0, text="Done!")

        with st.spinner("Generating Excel..."):
            excel_bytes = to_excel_bytes(all_results)
            log.info("Excel generated with %d sheets", len(all_results))

        st.download_button(
            label="Download Full Results (Excel — one sheet per tech stack)",
            data=excel_bytes,
            file_name="job_matches.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def tab_linkedin_search():
    st.header("LinkedIn Profile Finder")
    st.write("Paste a job requirement below and find matching LinkedIn profiles via Google search.")

    requirement = st.text_area(
        "Paste job requirement here",
        height=200,
        placeholder="e.g. Looking for a Java Spring Boot Developer with 5+ years experience in microservices, AWS, Kafka. Location: Charlotte, NC",
    )

    col1, col2 = st.columns(2)
    with col1:
        num_results = st.slider("Number of profiles to find", min_value=5, max_value=30, value=15)
    with col2:
        custom_location = st.text_input("Override location (optional)", "")

    if st.button("Search LinkedIn Profiles", type="primary"):
        if not requirement.strip():
            st.warning("Please paste a job requirement first.")
            return

        with st.spinner("Analyzing requirement..."):
            parsed = parse_requirement(requirement)

        if custom_location:
            parsed["location"] = custom_location

        st.subheader("Extracted from requirement")
        st.write(f"**Role:** {parsed['role'] or 'Not detected'}")
        st.write(f"**Skills:** {', '.join(parsed['skills']) or 'None detected'}")
        st.write(f"**Location:** {parsed['location'] or 'Not specified'}")

        queries = build_linkedin_queries(parsed["role"], parsed["skills"], parsed["location"])
        st.code(queries[0], language="text")

        # Always show a manual Google link as fallback
        manual_url = f"https://www.google.com/search?q={quote_plus(queries[0])}"
        st.markdown(f"[Open this search in Google manually]({manual_url})")

        with st.spinner("Searching Google for LinkedIn profiles (this may take 15-30 seconds)..."):
            profiles, used_query = search_linkedin_profiles(queries, num_results=num_results)

        if not profiles:
            st.warning(
                "No profiles found via automated search. Google may be rate-limiting. "
                "Use the manual Google link above to search directly."
            )
            return

        st.success(f"Found {len(profiles)} LinkedIn profile(s) using: `{used_query}`")

        for i, p in enumerate(profiles, 1):
            st.markdown(f"**{i}. [{p['name']}]({p['url']})**")

        excel_bytes = profiles_to_excel(profiles, used_query)
        st.download_button(
            label="Download Profiles (Excel)",
            data=excel_bytes,
            file_name="linkedin_profiles.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def main():
    st.set_page_config(page_title="Marketing Tool", page_icon="📄", layout="wide")
    st.title("Marketing Tool")

    tab1, tab2 = st.tabs(["Resume Matcher", "LinkedIn Profile Finder"])

    with tab1:
        tab_resume_matcher()

    with tab2:
        tab_linkedin_search()


if __name__ == "__main__":
    main()
